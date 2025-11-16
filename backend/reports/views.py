"""
Views for financial reports.
"""
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from datetime import datetime, date
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from weasyprint import HTML

from companies.models import Company
from transactions.accounting_service import AccountingService
from reports.trial_balance import TrialBalanceService


class ReportViewSet(viewsets.ViewSet):
    """ViewSet for financial reports."""
    
    permission_classes = []  # Temporarily allow any for testing
    
    def _get_company(self, request):
        """Get company from request."""
        company_id = request.query_params.get('company')
        if not company_id:
            # Try to get from user profile
            if hasattr(request.user, 'userprofile') and request.user.userprofile.active_company:
                return request.user.userprofile.active_company
            return None
        
        try:
            return Company.objects.get(id=company_id)
        except Company.DoesNotExist:
            return None
    
    def _parse_date(self, date_str):
        """Parse date string to date object."""
        if not date_str:
            return None
        try:
            return datetime.strptime(date_str, '%Y-%m-%d').date()
        except:
            return None
    
    @action(detail=False, methods=['get'])
    def balance_sheet(self, request):
        """Get balance sheet."""
        company = self._get_company(request)
        if not company:
            return Response(
                {'error': 'Company not found'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        end_date = self._parse_date(request.query_params.get('end_date'))
        export_format = request.query_params.get('export_format', '').lower()
        
        service = AccountingService(company)
        balance_sheet = service.get_balance_sheet(end_date=end_date)
        
        if export_format == 'excel':
            return self._export_balance_sheet_excel(balance_sheet, company)
        elif export_format == 'pdf':
            return self._export_balance_sheet_pdf(balance_sheet, company)
        elif export_format == 'csv':
            return self._export_balance_sheet_csv(balance_sheet, company)
        
        return Response(balance_sheet)
    
    @action(detail=False, methods=['get'])
    def income_statement(self, request):
        """Get income statement."""
        company = self._get_company(request)
        if not company:
            return Response(
                {'error': 'Company not found'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        start_date = self._parse_date(request.query_params.get('start_date'))
        end_date = self._parse_date(request.query_params.get('end_date'))
        export_format = request.query_params.get('export_format', '').lower()
        
        service = AccountingService(company)
        income_statement = service.get_income_statement(
            start_date=start_date,
            end_date=end_date
        )
        
        if export_format == 'excel':
            return self._export_income_statement_excel(income_statement, company)
        elif export_format == 'pdf':
            return self._export_income_statement_pdf(income_statement, company)
        elif export_format == 'csv':
            return self._export_income_statement_csv(income_statement, company)
        
        return Response(income_statement)
    
    @action(detail=False, methods=['get'])
    def trial_balance(self, request):
        """Get trial balance."""
        company = self._get_company(request)
        if not company:
            return Response(
                {'error': 'Company not found'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        start_date = self._parse_date(request.query_params.get('start_date'))
        end_date = self._parse_date(request.query_params.get('end_date'))
        
        trial_balance_data = TrialBalanceService.generate(
            company=company,
            start_date=start_date,
            end_date=end_date
        )
        
        return Response(trial_balance_data)
    
    @action(detail=False, methods=['get'])
    def cash_flow(self, request):
        """Get cash flow statement."""
        company = self._get_company(request)
        if not company:
            return Response(
                {'error': 'Company not found'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        start_date = self._parse_date(request.query_params.get('start_date'))
        end_date = self._parse_date(request.query_params.get('end_date'))
        
        export_format = request.query_params.get('export_format', '').lower()
        
        service = AccountingService(company)
        cash_flow = service.get_cash_flow_statement(
            start_date=start_date,
            end_date=end_date
        )
        
        if export_format == 'csv':
            return self._export_cash_flow_csv(cash_flow, company)
        
        return Response(cash_flow)
    
    def _export_balance_sheet_excel(self, data, company):
        """Export balance sheet to Excel."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Balance Sheet"
        
        # Header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        ws['A1'] = company.name
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = "Balance Sheet"
        ws['A2'].font = Font(bold=True, size=12)
        ws['A3'] = f"As of {data['date']}"
        
        row = 5
        
        # Assets
        ws[f'A{row}'] = "ASSETS"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        row += 1
        
        for item in data['assets']['items']:
            ws[f'A{row}'] = f"  {item['account_name']}"
            ws[f'B{row}'] = item['balance']
            ws[f'B{row}'].number_format = '$#,##0.00'
            row += 1
        
        ws[f'A{row}'] = "Total Assets"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = data['assets']['total']
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'B{row}'].font = Font(bold=True)
        row += 2
        
        # Liabilities
        ws[f'A{row}'] = "LIABILITIES"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        row += 1
        
        for item in data['liabilities']['items']:
            ws[f'A{row}'] = f"  {item['account_name']}"
            ws[f'B{row}'] = item['balance']
            ws[f'B{row}'].number_format = '$#,##0.00'
            row += 1
        
        ws[f'A{row}'] = "Total Liabilities"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = data['liabilities']['total']
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'B{row}'].font = Font(bold=True)
        row += 2
        
        # Equity
        ws[f'A{row}'] = "EQUITY"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        row += 1
        
        for item in data['equity']['items']:
            ws[f'A{row}'] = f"  {item['account_name']}"
            ws[f'B{row}'] = item['balance']
            ws[f'B{row}'].number_format = '$#,##0.00'
            row += 1
        
        ws[f'A{row}'] = "Total Equity"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = data['equity']['total']
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'B{row}'].font = Font(bold=True)
        row += 2
        
        # Total
        ws[f'A{row}'] = "TOTAL LIABILITIES & EQUITY"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = data['total_liabilities_and_equity']
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'B{row}'].font = Font(bold=True)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        
        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="balance_sheet_{data["date"]}.xlsx"'
        wb.save(response)
        
        return response
    
    def _export_income_statement_excel(self, data, company):
        """Export income statement to Excel."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Income Statement"
        
        # Header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        ws['A1'] = company.name
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = "Income Statement"
        ws['A2'].font = Font(bold=True, size=12)
        ws['A3'] = f"For the period {data['start_date']} to {data['end_date']}"
        
        row = 5
        
        # Revenue
        ws[f'A{row}'] = "REVENUE"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        row += 1
        
        for item in data['revenues']['items']:
            ws[f'A{row}'] = f"  {item['account_name']}"
            ws[f'B{row}'] = item['amount']
            ws[f'B{row}'].number_format = '$#,##0.00'
            row += 1
        
        ws[f'A{row}'] = "Total Revenue"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = data['revenues']['total']
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'B{row}'].font = Font(bold=True)
        row += 2
        
        # Expenses
        ws[f'A{row}'] = "EXPENSES"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        row += 1
        
        for item in data['expenses']['items']:
            ws[f'A{row}'] = f"  {item['account_name']}"
            ws[f'B{row}'] = item['amount']
            ws[f'B{row}'].number_format = '$#,##0.00'
            row += 1
        
        ws[f'A{row}'] = "Total Expenses"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = data['expenses']['total']
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'B{row}'].font = Font(bold=True)
        row += 2
        
        # Net Income
        ws[f'A{row}'] = "NET INCOME"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'B{row}'] = data['net_income']
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'B{row}'].font = Font(bold=True, size=12)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        
        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="income_statement_{data["start_date"]}_{data["end_date"]}.xlsx"'
        wb.save(response)
        
        return response
    
    def _export_balance_sheet_pdf(self, data, company):
        """Export balance sheet to PDF."""
        html_content = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; }}
                h1 {{ text-align: center; }}
                h2 {{ text-align: center; color: #333; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
                th {{ background-color: #4472C4; color: white; padding: 10px; text-align: left; }}
                td {{ padding: 8px; border-bottom: 1px solid #ddd; }}
                .total {{ font-weight: bold; background-color: #f0f0f0; }}
                .amount {{ text-align: right; }}
            </style>
        </head>
        <body>
            <h1>{company.name}</h1>
            <h2>Balance Sheet</h2>
            <p style="text-align: center;">As of {data['date']}</p>
            
            <h3>ASSETS</h3>
            <table>
                {''.join([f'<tr><td>{item["account_name"]}</td><td class="amount">${item["balance"]:,.2f}</td></tr>' for item in data['assets']['items']])}
                <tr class="total"><td>Total Assets</td><td class="amount">${data['assets']['total']:,.2f}</td></tr>
            </table>
            
            <h3>LIABILITIES</h3>
            <table>
                {''.join([f'<tr><td>{item["account_name"]}</td><td class="amount">${item["balance"]:,.2f}</td></tr>' for item in data['liabilities']['items']])}
                <tr class="total"><td>Total Liabilities</td><td class="amount">${data['liabilities']['total']:,.2f}</td></tr>
            </table>
            
            <h3>EQUITY</h3>
            <table>
                {''.join([f'<tr><td>{item["account_name"]}</td><td class="amount">${item["balance"]:,.2f}</td></tr>' for item in data['equity']['items']])}
                <tr class="total"><td>Total Equity</td><td class="amount">${data['equity']['total']:,.2f}</td></tr>
            </table>
            
            <h3 style="margin-top: 30px;">TOTAL LIABILITIES & EQUITY: ${data['total_liabilities_and_equity']:,.2f}</h3>
        </body>
        </html>
        """
        
        pdf = HTML(string=html_content).write_pdf()
        
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="balance_sheet_{data["date"]}.pdf"'
        
        return response
    
    def _export_income_statement_pdf(self, data, company):
        """Export income statement to PDF."""
        html_content = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; }}
                h1 {{ text-align: center; }}
                h2 {{ text-align: center; color: #333; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
                th {{ background-color: #4472C4; color: white; padding: 10px; text-align: left; }}
                td {{ padding: 8px; border-bottom: 1px solid #ddd; }}
                .total {{ font-weight: bold; background-color: #f0f0f0; }}
                .amount {{ text-align: right; }}
                .net-income {{ font-size: 18px; margin-top: 30px; text-align: center; }}
            </style>
        </head>
        <body>
            <h1>{company.name}</h1>
            <h2>Income Statement</h2>
            <p style="text-align: center;">For the period {data['start_date']} to {data['end_date']}</p>
            
            <h3>REVENUE</h3>
            <table>
                {''.join([f'<tr><td>{item["account_name"]}</td><td class="amount">${item["amount"]:,.2f}</td></tr>' for item in data['revenues']['items']])}
                <tr class="total"><td>Total Revenue</td><td class="amount">${data['revenues']['total']:,.2f}</td></tr>
            </table>
            
            <h3>EXPENSES</h3>
            <table>
                {''.join([f'<tr><td>{item["account_name"]}</td><td class="amount">${item["amount"]:,.2f}</td></tr>' for item in data['expenses']['items']])}
                <tr class="total"><td>Total Expenses</td><td class="amount">${data['expenses']['total']:,.2f}</td></tr>
            </table>
            
            <h3 class="net-income">NET INCOME: ${data['net_income']:,.2f}</h3>
        </body>
        </html>
        """
        
        pdf = HTML(string=html_content).write_pdf()
        
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="income_statement_{data["start_date"]}_{data["end_date"]}.pdf"'
        
        return response



    def _export_balance_sheet_csv(self, data, company):
        """Export balance sheet to CSV."""
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([company.name])
        writer.writerow(['Balance Sheet'])
        writer.writerow([f'As of {data["date"]}'])
        writer.writerow([])
        
        # Assets
        writer.writerow(['ASSETS'])
        writer.writerow(['Account', 'Balance'])
        for item in data['assets']['items']:
            writer.writerow([item['account_name'], f"${item['balance']:,.2f}"])
        writer.writerow(['Total Assets', f"${data['assets']['total']:,.2f}"])
        writer.writerow([])
        
        # Liabilities
        writer.writerow(['LIABILITIES'])
        writer.writerow(['Account', 'Balance'])
        for item in data['liabilities']['items']:
            writer.writerow([item['account_name'], f"${item['balance']:,.2f}"])
        writer.writerow(['Total Liabilities', f"${data['liabilities']['total']:,.2f}"])
        writer.writerow([])
        
        # Equity
        writer.writerow(['EQUITY'])
        writer.writerow(['Account', 'Balance'])
        for item in data['equity']['items']:
            writer.writerow([item['account_name'], f"${item['balance']:,.2f}"])
        writer.writerow(['Total Equity', f"${data['equity']['total']:,.2f}"])
        writer.writerow([])
        
        # Total
        writer.writerow(['TOTAL LIABILITIES & EQUITY', f"${data['total_liabilities_and_equity']:,.2f}"])
        
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="balance_sheet_{data["date"]}.csv"'
        
        return response
    
    def _export_income_statement_csv(self, data, company):
        """Export income statement to CSV."""
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([company.name])
        writer.writerow(['Income Statement'])
        writer.writerow([f'For the period {data["start_date"]} to {data["end_date"]}'])
        writer.writerow([])
        
        # Revenue
        writer.writerow(['REVENUE'])
        writer.writerow(['Account', 'Amount'])
        for item in data['revenues']['items']:
            writer.writerow([item['account_name'], f"${item['amount']:,.2f}"])
        writer.writerow(['Total Revenue', f"${data['revenues']['total']:,.2f}"])
        writer.writerow([])
        
        # Expenses
        writer.writerow(['EXPENSES'])
        writer.writerow(['Account', 'Amount'])
        for item in data['expenses']['items']:
            writer.writerow([item['account_name'], f"${item['amount']:,.2f}"])
        writer.writerow(['Total Expenses', f"${data['expenses']['total']:,.2f}"])
        writer.writerow([])
        
        # Net Income
        writer.writerow(['NET INCOME', f"${data['net_income']:,.2f}"])
        
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="income_statement_{data["start_date"]}_{data["end_date"]}.csv"'
        
        return response
    
    def _export_cash_flow_csv(self, data, company):
        """Export cash flow statement to CSV."""
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([company.name])
        writer.writerow(['Cash Flow Statement'])
        writer.writerow([f'For the period {data["start_date"]} to {data["end_date"]}'])
        writer.writerow([])
        
        # Cash flow details
        writer.writerow(['Beginning Cash Balance', f"${data['beginning_cash_balance']:,.2f}"])
        writer.writerow([])
        writer.writerow(['Operating Activities', f"${data['operating_activities']:,.2f}"])
        writer.writerow(['Investing Activities', f"${data['investing_activities']:,.2f}"])
        writer.writerow(['Financing Activities', f"${data['financing_activities']:,.2f}"])
        writer.writerow([])
        writer.writerow(['Net Change in Cash', f"${data['net_change_in_cash']:,.2f}"])
        writer.writerow([])
        writer.writerow(['Ending Cash Balance', f"${data['ending_cash_balance']:,.2f}"])
        
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="cash_flow_{data["start_date"]}_{data["end_date"]}.csv"'
        
        return response

